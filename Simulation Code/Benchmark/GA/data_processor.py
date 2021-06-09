import json
import csv
import openpyxl


def load_xls(xls_path, sheet_name):
    wb = openpyxl.load_workbook(xls_path)
    ws = wb[sheet_name]
    return ws


def get_slots(ws):
    ans = {}
    for row in range(2, ws.max_row-1):
        if not ws.cell(row=row, column=1).value == None:
            key = ws.cell(row=row, column=1).value
            slots = ws.cell(row=row, column=2).value
            ans[key] = slots
    return ans


def get_location(ws):
    print(ws.max_row)
    ans = {}
    for row in range(2, ws.max_row+1):
        print(ws.cell(row=row, column=1).value)
        if not ws.cell(row=row, column=1).value == None:
            key = ws.cell(row=row, column=1).value
            location = ws.cell(row=row, column=2).value
            ans[key] = location
    return ans


def get_bandwidth(ws):
    ans = []
    for row in range(2, ws.max_row):
        tmp = []
        for col in range(2, ws.max_column):
            if not ws.cell(row=row, column=col).value == '-':
                tmp.append(ws.cell(row=row, column=col).value)
            else:
                tmp.append(0)
        ans.append(tmp)
    for k in range(len(tmp)):
        for i in range(len(tmp)):
            for j in range(len(tmp)):
                if ans[i][j] < min(ans[i][k], ans[k][j]):
                    ans[i][j] = min(ans[i][k], ans[k][j])
    return ans


def get_data_req(ws):
    ans = {}
    for row in range(2, ws.max_row+1):
        for col in range(2, ws.max_column+1):
            f = ws.cell(row=row, column=1).value
            t = ws.cell(row=1, column=col).value
            if not f in ans.keys():
                ans[f] = {}
            if not ws.cell(row=row, column=col).value == None:
                ans[f][t] = ws.cell(row=row, column=col).value
    return ans


# ws = load_xls('./ToyData.xlsx', 'Data Center Details')
# slot_dict = get_slots(ws)
# json.dump(slot_dict, open('slot.json', 'w'))

# ws = load_xls('./ToyData.xlsx', 'Sheet1')
# loc_dict = get_location(ws)
# json.dump(loc_dict, open('./data/location.json', 'w'))

ws = load_xls('./ToyData.xlsx', 'Inter-Datacenter Links')
band_width_list = get_bandwidth(ws)
json.dump(band_width_list, open('./data/bandwidth.json', 'w'))

# ws = load_xls('./ToyData.xlsx', 'Sheet2')
# data_req = get_data_req(ws)
# json.dump(data_req, open('./data/data_req.json', 'w'))
